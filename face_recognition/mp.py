import face_recognition
import cv2
import numpy as np
import os
from datetime import datetime 
import tkinter as tk
from PIL import Image, ImageTk
import openpyxl

# Function to update the GUI display
def update_gui(name):
    attendance_label.config(text=f"{name} Present")

# Function to update the camera feed in the GUI
def update_camera_feed(known_face_encodings, known_face_names):
    _, frame = video_capture.read()
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    frame = cv2.resize(frame, (width, height))

    # Find face locations and encodings
    face_locations = face_recognition.face_locations(frame)
    face_encodings = face_recognition.face_encodings(frame, face_locations)

    # Loop through each face
    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"

        if True in matches:
            name = known_face_names[matches.index(True)]
            update_gui(name)

            # Update attendance record
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            update_attendance(name, current_time)

            # Draw a green rectangle around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        else:
            # Draw a red rectangle around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (255, 0, 0), 2)

    # Convert the frame to PIL format
    img = Image.fromarray(frame)
    imgtk = ImageTk.PhotoImage(image=img)

    # Update the label with the new frame
    camera_label.imgtk = imgtk
    camera_label.configure(image=imgtk)
    if not stop_camera:
        camera_label.after(10, update_camera_feed, known_face_encodings, known_face_names)  # Update every 10 milliseconds

# Function to stop the camera and close the window
def stop_camera():
    global stop_camera
    stop_camera = True
    video_capture.release()
    root.destroy()

# Initialize Tkinter
root = tk.Tk()
root.title("Attendance System")

# Title label
title_label = tk.Label(root, text="Face Recognition Attendance System", font=('Helvetica', 20, 'bold'))
title_label.pack(pady=10)

# Date and time label
current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
datetime_label = tk.Label(root, text=current_datetime, font=('Helvetica', 14))
datetime_label.pack()

# Create a label for attendance display
attendance_label = tk.Label(root, text="", font=('Helvetica', 18))
attendance_label.pack(pady=20)

# Create a label for camera feed
camera_label = tk.Label(root)
camera_label.pack()

# Button to stop the camera and close the window
stop_button = tk.Button(root, text="Stop and Exit", command=stop_camera, font=('Helvetica', 14))
stop_button.pack(pady=10)

# Initialize video capture
video_capture = cv2.VideoCapture(0)

# Get the initial frame dimensions
width = int(video_capture.get(cv2.CAP_PROP_FRAME_WIDTH))
height = int(video_capture.get(cv2.CAP_PROP_FRAME_HEIGHT))

# Load known face encodings and names
known_face_encodings = []
known_face_names = []
for filename in os.listdir("photos"):
    if filename.endswith(".jpg"):
        name = os.path.splitext(filename)[0]
        image = face_recognition.load_image_file("photos/" + filename)
        encoding = face_recognition.face_encodings(image)[0]
        known_face_encodings.append(encoding)
        known_face_names.append(name)

# Create attendance Excel file if not exists
def create_attendance_file():
    if not os.path.exists('attendance.xlsx'):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Name', 'P/A', 'Time'])  # Add header row
        wb.save('attendance.xlsx')

# Function to update the attendance record
def update_attendance(name, current_time):
    create_attendance_file()
    wb = openpyxl.load_workbook('attendance.xlsx')
    sheet = wb.active
    student_names = load_student_names()

    if name in student_names:
        # Find the row index of the student
        row_index = student_names.index(name) + 2  # Adding 2 to account for header row and 0-based indexing

        # Update the attendance for the student
        sheet.cell(row=row_index, column=2).value = "P"  # Mark present in the P/A column
        sheet.cell(row=row_index, column=3).value = current_time  # Update the time column
    else:
        # Append a new entry for the student
        sheet.append([name, "A", current_time])  # Append name, "Absent", and time

    # Save the updated workbook
    wb.save('attendance.xlsx')

# Function to load student names from Excel file
def load_student_names():
    student_names = []
    if os.path.exists('attendance.xlsx'):
        wb = openpyxl.load_workbook('attendance.xlsx')
        sheet = wb.active
        student_names = [cell.value for cell in sheet['A'][1:]]
        wb.close()
    return student_names


# Flag to stop camera
stop_camera = False

# Main loop
update_camera_feed(known_face_encodings, known_face_names)
root.mainloop()
